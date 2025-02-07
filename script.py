import os
import time
import pyautogui
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.chrome.options import Options
import datetime
from selenium.webdriver.common.keys import Keys


download_path = r"C:\Users\DalalD\Downloads"
existing_file_path = r"C:\Users\DalalD\Desktop\ICM_OUTPUT.xlsx"

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(10)

# Opening Page
driver.get("https://hnicorporation.sharepoint.com/sites/hnimdmitem/Lists/Oracle%20Part%20CreationUpdate/MDM%20Item.aspx")

try:
    # Login Page
    mail = WebDriverWait(driver, 30).until(expected_conditions.element_to_be_clickable((By.NAME, "loginfmt")))
    mail.send_keys('dalald@hni-asia.com')

    # Next button
    next_button = WebDriverWait(driver, 30).until(expected_conditions.element_to_be_clickable((By.CSS_SELECTOR, "#idSIButton9")))
    next_button.click()
    print("Mail entered")

    # Office 365
    submit_button = WebDriverWait(driver, 30).until(
        expected_conditions.element_to_be_clickable((By.XPATH, "//input[@class='sa-2019-button']")))
    submit_button.click()

    # Credentials
    WebDriverWait(driver, 30).until(expected_conditions.presence_of_element_located((By.XPATH, "//input[@type='password']")))
    driver.find_element(By.XPATH, "//input[@type='password']").send_keys('Graduateengineertrainee@25')
    driver.find_element(By.CSS_SELECTOR, ".btn").click()
    print("Password Entered")

    # Stay signed in page
    Yes_btn = WebDriverWait(driver, 30).until(expected_conditions.element_to_be_clickable((By.CSS_SELECTOR, "#idSIButton9")))
    Yes_btn.click()

    # Table & rows
    table = WebDriverWait(driver, 30).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, 'ms-listviewtable')))
    rows = table.find_elements(By.TAG_NAME, 'tr')

    # Today's date
    today = datetime.date.today()

    # Dates for today, yesterday, and the day before yesterday
    dates_to_check = [(today - datetime.timedelta(days=1)).strftime('%#m/%#d/%Y')]

    print(f"Dates to check: {dates_to_check}")

    # Getting IDs for today and the last two days
    IDs = []

    for row in rows:
        try:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) > 23:
                icm = cols[1].text
                date_value = cols[23].text

                try:
                    date_only = datetime.datetime.strptime(date_value, "%m/%d/%Y %I:%M %p").strftime("%#m/%#d/%Y")
                except ValueError:
                    print(f"Skipping invalid date format: {date_value},{icm}")
                    continue  # Skip this row if date format is incorrect

                if any(date in date_value for date in dates_to_check):
                    IDs.append((icm,date_value))

        except Exception as e:
            print(f"Error processing row: {e}")
            continue

    # print(f"IDs for the last three days: count({IDs})") 
    print(len(IDs))
    driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")


    for i, date_created in IDs:
        try:
            print(f"Searching for ID: {i}")

            # Searching ID
            search_id = WebDriverWait(driver, 30).until(expected_conditions.presence_of_element_located((By.ID, 'inplaceSearchDiv_WPQ4_lsinput')))
            search_id.clear()
            search_id.send_keys(i)
            search_id.send_keys(Keys.ENTER)

            # Clicking the title to go to the frame
            WebDriverWait(driver, 30).until(expected_conditions.invisibility_of_element((By.CLASS_NAME, "ms-dlgOverlay")))
            click_title = WebDriverWait(driver, 30).until(expected_conditions.element_to_be_clickable((By.CLASS_NAME, "ms-listlink")))
            click_title.click()
            print(f"Clicked on title: {i}")

            # Entering the frame
            iframes = WebDriverWait(driver, 30).until(expected_conditions.presence_of_all_elements_located((By.TAG_NAME, 'iframe')))
            print(f"Total iframes found: {len(iframes)}")
            driver.switch_to.frame(iframes[5])
            print("Found the frame")

            # Scrolling
            driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")

            try:
                # Download files
                download_files = driver.find_element(By.ID, "idAttachmentsTable")
                multiple_attachments = download_files.find_elements(By.TAG_NAME, 'tr')

                if not multiple_attachments:
                    raise Exception("No attachments found")

                for index, attachment in enumerate(multiple_attachments):
                    action = ActionChains(driver)
                    action.context_click(attachment).perform()
                    time.sleep(1)
                    # Mouse actions, right-click and select the fourth option
                    pyautogui.press('down', presses=4)
                    pyautogui.press('enter')
                    time.sleep(2)
                    # Save the file with the ID name
                    file_name = f"{i}.xlsx" if index == 0 else f"{i}_{index}.xlsx"
                    pyautogui.write(file_name)
                    pyautogui.press('enter')
                    time.sleep(2)

                    # Full path to the downloaded file
                    downloaded_file_path = os.path.join(download_path, file_name)

                    time.sleep(20)

                    if os.path.exists(downloaded_file_path):
                        print("File found.")
                    else:
                        print("File not found.")

                    file_downloaded = load_workbook(downloaded_file_path)
                    sheet_active = file_downloaded.active

                    mst_org_row = None
                    for row in sheet_active.iter_rows(min_row=1, max_row=10):
                        row_values = [cell.value for cell in row]
                        if "MST" in row_values and "Org" in row_values:
                            mst_org_row = row[0].row  # Row number
                            break

                    if mst_org_row is None:
                        raise Exception("Row containing both 'MST' and 'Org' not found.")

                    headers = {}
                    for row in sheet_active.iter_rows(min_row=mst_org_row, max_row=10):
                        for cell in row:
                            if cell.value in ["Item Number", "Alias Item Number", "Base Model Core Assy Item #", "Alias", "Item Number (OSFG Family Items cannot have more than 30 characters)", "Organization Code", "Org"]:
                                headers[cell.value] = cell.column

                    if not headers:
                        raise Exception("Required headers not found.")

                    # List of headers
                    item_number_headers = ["Item Number", "Alias Item Number", "Base Model Core Assy Item #", "Alias", "Item Number (OSFG Family Items cannot have more than 30 characters)"]
                    org_code_headers = ["Organization Code", "Org"]

                    # Finding headers
                    item_number_col = next((headers[header] for header in item_number_headers if header in headers),None)
                    org_code_col = next((headers[header] for header in org_code_headers if header in headers), None)

                    if item_number_col is None or org_code_col is None:
                        raise Exception("Required headers not found.")

                    # Extract data starting from two rows after the "MST" and "Org" row
                    data_to_append = []
                    for row in sheet_active.iter_rows(min_row=mst_org_row + 2, max_row=sheet_active.max_row):
                        item_number = row[item_number_col - 1].value
                        org_code = row[org_code_col - 1].value
                        if item_number:
                            data_to_append.append((item_number, org_code if org_code else "", date_created))

                    # Open the existing Excel file to append data
                    file_existing = load_workbook(existing_file_path)
                    sheet_existing = file_existing.active

                    next_row = sheet_existing.max_row + 1

                    # Append each data row to the existing file
                    for item_number, org_code, date_created in data_to_append:
                        sheet_existing.cell(row=next_row, column=1, value=item_number)
                        sheet_existing.cell(row=next_row, column=2, value=org_code)
                        sheet_existing.cell(row=next_row, column=3, value=i)
                        sheet_existing.cell(row=next_row, column=4, value=date_created)
                        next_row += 1

                    file_existing.save(existing_file_path)
                    file_existing.close()
                    file_downloaded.close()
                    print(f"Data from {downloaded_file_path} appended successfully.")

            except Exception as e:
                file_existing = load_workbook(existing_file_path)
                sheet_existing = file_existing.active

                next_row = sheet_existing.max_row + 1

                sheet_existing.cell(row=next_row, column=1, value="N/A")
                sheet_existing.cell(row=next_row, column=2, value="N/A")
                sheet_existing.cell(row=next_row, column=3, value=i)
                sheet_existing.cell(row=next_row, column=4, value= date_created)
                sheet_existing.cell(row=next_row, column=5, value=str(e))

                file_existing.save(existing_file_path)
                file_existing.close()
                print(f"Error message for ID {i}: {e}")

            # Refresh and switch to the default window
            driver.refresh()
            driver.switch_to.default_content()
            print("Switched back to default content")

        except Exception as e:
            print(f"Error: {e}")

except Exception as e:
    print(f"Error encountered: {e}")
